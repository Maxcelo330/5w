# app_mejorada.py
import streamlit as st
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
import io
import qrcode  # Se necesita: pip install qrcode
from PIL import Image # Se necesita: pip install pillow
import socket # Para obtener la IP local

# --- Función para encontrar la ruta de recursos (útil para la plantilla) ---
def resource_path(relative_path):
    """ Obtiene la ruta absoluta al recurso, funciona para desarrollo y para PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- Función para obtener la IP local de la red ---
def get_local_ip():
    """
    Intenta obtener la dirección IP local de la máquina en la red.
    """
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # No necesita ser alcanzable
        s.connect(('10.255.255.255', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1' # Fallback a localhost si no se puede determinar la IP
    finally:
        s.close()
    return IP

# --- Configuración de la página ---
st.set_page_config(page_title="Análisis 5 Porqués", page_icon="📝", layout="wide")

# --- TÍTULO ---
st.title("📝 Formulario de Análisis - 5 Porqués")
st.markdown("Completa todos los campos del formulario para generar el reporte en Excel basado en tu plantilla.")
st.divider()

# --- SECCIÓN 1: INFORMACIÓN GENERAL ---
st.header("1. Información General del Evento")
col1, col2, col3 = st.columns(3)
with col1:
    fecha_evento = st.date_input("Fecha del Evento")
    sector = st.text_input("Sector")
    operacion = st.text_input("Operación")
with col2:
    periodo_tiempo = st.text_input("Tiempo de Parada / Periodo de Tiempo")
    equipo_proceso = st.text_input("Equipo o Proceso")
    disparador = st.text_input("Disparador")
with col3:
    turno = st.selectbox("Turno", ["Turno A", "Turno B", "Turno C", "Otro"])

st.divider()

# --- SECCIÓN 2: DESCRIPCIÓN DEL PROBLEMA ---
st.header("2. Descripción del Problema")
col_prob_1, col_prob_2 = st.columns(2)
with col_prob_1:
    problema_sintomas = st.text_area("¿Cuál fue el problema? ¿Qué pasó? ¿Cuáles fueron los síntomas?", height=150)
with col_prob_2:
    acciones_realizadas = st.text_area("¿Qué acciones se realizaron? ¿Hay acciones contingentes?", height=150)

st.divider()

# --- SECCIÓN 3: ANÁLISIS 5 PORQUÉS ---
st.header("3. Análisis de Causa Raíz (5 Porqués)")
porque1 = st.text_input("1. ¿Por qué?")
porque2 = st.text_input("2. ¿Por qué?")
porque3 = st.text_input("3. ¿Por qué?")
porque4 = st.text_input("4. ¿Por qué?")
porque5 = st.text_input("5. ¿Por qué?")

st.divider()

# --- SECCIÓN 4: PARTICIPANTES Y PREVENCIÓN ---
st.header("4. Participantes y Prevención")
col_part_1, col_part_2 = st.columns(2)
with col_part_1:
    detectores = st.text_area("Detectores (¿Quiénes participaron?)", height=100)
with col_part_2:
    reparadores = st.text_area("Reparadores (¿Quiénes participaron?)", height=100)

prevencion = st.text_area("¿Cómo crees que se podría evitar que vuelva a ocurrir?", height=150)

st.divider()

# --- SECCIÓN 5: SEGUIMIENTO Y PLAN DE ACCIÓN ---
st.header("5. Seguimiento y Plan de Acción")
col_seg_1, col_seg_2 = st.columns(2)
with col_seg_1:
    seguimiento_adf = st.radio("¿Es necesario seguir con el análisis (ADF)?", ["Sí", "No"])
with col_seg_2:
    identifico_causa = st.radio("¿Se identificó la causa raíz, se necesitan terminar acciones?", ["Sí", "No"])

st.subheader("Plan de Acción (Obligatorio)")

# Inicializar el plan de acción en el estado de la sesión si no existe
if 'plan_de_accion' not in st.session_state:
    st.session_state.plan_de_accion = [{
        'id': 0, 'fecha': datetime.today(), 'actividad': '', 'responsable': '', 'fecha_cierre': datetime.today(), 'estado': ''
    }]

# Mostrar filas del plan de acción
rows_to_delete = []
for i, item in enumerate(st.session_state.plan_de_accion):
    cols = st.columns([2, 4, 3, 2, 2, 1])
    item['fecha'] = cols[0].date_input("Fecha", value=item['fecha'], key=f"pa_fecha_{item['id']}")
    item['actividad'] = cols[1].text_input("Actividad", value=item['actividad'], key=f"pa_actividad_{item['id']}")
    item['responsable'] = cols[2].text_input("Responsable", value=item['responsable'], key=f"pa_responsable_{item['id']}")
    item['fecha_cierre'] = cols[3].date_input("Fecha Cierre", value=item['fecha_cierre'], key=f"pa_fecha_cierre_{item['id']}")
    item['estado'] = cols[4].text_input("Estado", value=item['estado'], key=f"pa_estado_{item['id']}")
    if cols[5].button("🗑️", key=f"delete_{item['id']}"):
        if len(st.session_state.plan_de_accion) > 1:
            rows_to_delete.append(i)

# Eliminar las filas marcadas
for index in sorted(rows_to_delete, reverse=True):
    del st.session_state.plan_de_accion[index]
    st.rerun()

# Botón para añadir nueva fila
if st.button("➕ Añadir Actividad al Plan"):
    new_id = st.session_state.plan_de_accion[-1]['id'] + 1 if st.session_state.plan_de_accion else 0
    st.session_state.plan_de_accion.append({
        'id': new_id, 'fecha': datetime.today(), 'actividad': '', 'responsable': '', 'fecha_cierre': datetime.today(), 'estado': ''
    })
    st.rerun()

st.divider()

observaciones = st.text_area("Observaciones Generales", height=150)

st.divider()

# --- BOTÓN PARA GENERAR EL REPORTE ---
if st.button("🚀 Generar Reporte de Análisis", type="primary"):
    
    # Recolectar todas las respuestas en un diccionario
    answers = {
        # Info General
        "fecha_evento": fecha_evento.strftime("%Y-%m-%d"), "periodo_tiempo": periodo_tiempo, "turno": turno,
        "sector": sector, "equipo_proceso": equipo_proceso, "operacion": operacion, "disparador": disparador,
        # Problema
        "problema_sintomas": problema_sintomas, "acciones_realizadas": acciones_realizadas,
        # 5 Porqués
        "porque1": porque1, "porque2": porque2, "porque3": porque3, "porque4": porque4, "porque5": porque5,
        # Participantes y Prevención
        "detectores": detectores, "reparadores": reparadores, "prevencion": prevencion,
        # Seguimiento
        "seguimiento_adf": seguimiento_adf, "identifico_causa": identifico_causa,
        # Plan de Acción
        "plan_de_accion": st.session_state.plan_de_accion,
        # Observaciones
        "observaciones": observaciones
    }

    # --- Generación del archivo en memoria ---
    output = io.BytesIO()
    try:
        nombre_plantilla = "PLANTILLA ANALISIS 5 PORQUE.xlsx"
        plantilla_path = resource_path(os.path.join("assets", nombre_plantilla))
        
        if os.path.exists(plantilla_path):
            st.info(f"Usando la plantilla '{nombre_plantilla}' encontrada.")
            wb = load_workbook(plantilla_path)
            ws = wb.active
            
            # --- MAPEO DE RESPUESTAS A CELDAS (según tu última corrección) ---
            ws["E3"] = answers["fecha_evento"]
            ws["G3"] = answers["turno"]
            ws["F3"] = answers["periodo_tiempo"]
            ws["A6"] = answers["sector"]
            ws["D6"] = answers["equipo_proceso"]
            ws["A9"] = answers["operacion"]
            ws["D8"] = answers["disparador"]
            ws["A12"] = answers["problema_sintomas"]
            ws["F12"] = answers["acciones_realizadas"]
            ws["C14"] = answers["porque1"]
            ws["C15"] = answers["porque2"]
            ws["C16"] = answers["porque3"]
            ws["C17"] = answers["porque4"]
            ws["C18"] = answers["porque5"]
            ws["A22"] = answers["detectores"]
            ws["F22"] = answers["reparadores"]
            ws["A23"] = answers["prevencion"]
            ws["D28"] = answers["seguimiento_adf"]
            ws["H28"] = answers["identifico_causa"]

            start_row = 31 # Fila inicial en la plantilla de Excel
            for i, item in enumerate(answers["plan_de_accion"]):
                current_row = start_row + i
                ws[f"A{current_row}"] = item["fecha"].strftime("%Y-%m-%d")
                ws[f"B{current_row}"] = item["actividad"]
                ws[f"F{current_row}"] = item["responsable"]
                ws[f"H{current_row}"] = item["fecha_cierre"].strftime("%Y-%m-%d")
                ws[f"I{current_row}"] = item["estado"]
            
            ws["A36"] = answers["observaciones"]

            # --- Guardar y preparar para descarga ---
            wb.save(output)
            excel_data = output.getvalue()
            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_name = f"Analisis_5W_{now}.xlsx"
            
            st.success("¡Reporte generado con éxito!")
            st.download_button(
                label="📥 Descargar Reporte en Excel",
                data=excel_data,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error(f"Error: No se encontró el archivo '{nombre_plantilla}' en la carpeta 'assets'. Asegúrate de que exista.")

    except Exception as e:
        st.error(f"Ocurrió un error al generar el archivo: {e}")

st.divider()

# --- NUEVA SECCIÓN: CÓDIGO QR PARA ACCESO MÓVIL ---
st.header("📲 Acceso desde el Celular")
st.write("Escanea este código QR con tu celular para abrir esta página y llenar el formulario directamente.")

try:
    # Generar la URL con la IP de la red local
    local_ip = get_local_ip()
    port = "8501" # Puerto por defecto de Streamlit
    url = f"http://{local_ip}:{port}"

    # Generar la imagen del QR
    qr_img = qrcode.make(url)
    
    # Convertir a un formato que Streamlit pueda mostrar
    img_buffer = io.BytesIO()
    qr_img.save(img_buffer, "PNG")
    
    # Mostrar la URL y el QR
    st.write(f"O abre esta URL en tu navegador: **{url}**")
    st.image(img_buffer)
    st.info("Asegúrate de que tu celular esté conectado a la misma red Wi-Fi que esta computadora.")

except Exception as e:
    st.error(f"No se pudo generar el código QR. Error: {e}")
    st.warning("Asegúrate de estar conectado a una red.")

